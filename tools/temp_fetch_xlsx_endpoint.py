import os
import requests
import google.auth.transport.requests
from google_drive_tool import GoogleDriveTool
import dotenv
import openpyxl
from io import BytesIO

dotenv.load_dotenv()

def debug():
    tool = GoogleDriveTool()
    root_folder = os.getenv("GOOGLE_DRIVE_ROOT_ID")
    
    farms = tool.get_farm_folders(root_folder)
    target_id = None
    for farm in farms:
        for file in tool.list_xlsx_files(farm['id']):
            name_upper = file['name'].upper()
            if "1A" in name_upper and "BBK" in name_upper:
                target_id = file['id']
                break
        if target_id: break
        
    # Get auth token
    req = google.auth.transport.requests.Request()
    tool.creds.refresh(req)
    token = tool.creds.token
    
    url = f"https://docs.google.com/spreadsheets/d/{target_id}/export?format=xlsx"
    print(f"Fetching from: {url}")
    
    headers = {"Authorization": f"Bearer {token}"}
    resp = requests.get(url, headers=headers)
    
    if resp.status_code == 200:
        print("Success! Loading into openpyxl...")
        content = BytesIO(resp.content)
        wb = openpyxl.load_workbook(content, data_only=True, read_only=True)
        
        # Look for Data Harian
        for name in wb.sheetnames:
            if 'Harian' in name:
                print(f"Found sheet: {name}")
                ws = wb[name]
                rows = list(ws.iter_rows(values_only=True))
                
                print("--- Checking Row 275-290 (Mar 2026) ---")
                for r_idx in range(275, 290):
                    if r_idx < len(rows):
                        print(f"Row {r_idx+1}: {rows[r_idx][:10]}")
        wb.close()
    else:
        print(f"Error Body: {resp.text[:500]}")

if __name__ == "__main__":
    debug()
