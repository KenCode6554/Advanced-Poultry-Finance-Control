import sys, os, io
sys.path.append(os.getcwd())
from tools.google_drive_tool import GoogleDriveTool
import openpyxl, datetime

tool = GoogleDriveTool()

# Sample files: Jantan JTP, KD 5 JTP, KD 7 BBK
samples = [
    ('REC KD Jantan JTP Mojogedang.xlsx',  '1_AVPCyEuVLUulSeUzPYehCqkpZQFrRw0'),
    ('REC KD 5 PL241P JTP Mojogedang .xlsx', '1IaVF9iIXFfREX1V6d6yzPugYhVoSfEwF'),
    ('Rec P. fajar kd 7A BBK.xlsx', None),
]

for fname, fid in samples:
    print(f'\n=== {fname} ===')
    if not fid:
        results = tool.drive_service.files().list(q=f"name='{fname}'", spaces='drive').execute().get('files', [])
        if results: fid = results[0]['id']
    if not fid:
        print('  NOT FOUND'); continue

    content = tool.download_file(fid)
    wb = openpyxl.load_workbook(io.BytesIO(content.getvalue()), data_only=True)
    harian_sheets = [s for s in wb.sheetnames if 'harian' in s.lower()]
    sheet = wb[harian_sheets[0]] if harian_sheets else wb.worksheets[0]
    print(f'  Sheet: {sheet.title}')

    # Print rows 4-9 to see header structure
    for r in range(4, 10):
        vals = [str(sheet.cell(row=r, column=c).value or '').strip() for c in range(1, 20)]
        non_empty = [(i+1, v) for i, v in enumerate(vals) if v]
        if non_empty:
            print(f'  Row {r}: {non_empty}')
