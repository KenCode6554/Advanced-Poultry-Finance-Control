import sys, os, io
sys.path.append(os.getcwd())
from tools.google_drive_tool import GoogleDriveTool
tool = GoogleDriveTool()

fid = None
for q in tool.drive_service.files().list(q="name = 'REC KD Jantan JTP Mojogedang.xlsx'", spaces='drive').execute().get('files', []):
    fid = q['id']

if fid:
    print('Downloading', fid)
    content = tool.download_file(fid)
    import openpyxl, datetime
    wb = openpyxl.load_workbook(io.BytesIO(content.getvalue()), data_only=True)
    sheet_name = [s for s in wb.sheetnames if 'harian' in s.lower()]
    sheet = wb[sheet_name[0]] if sheet_name else wb.worksheets[0]
    
    col_tgl = col_mati = col_afkir = None
    for c in range(1, 40):
        v_1 = str(sheet.cell(row=4, column=c).value or '').lower()
        v_2 = str(sheet.cell(row=5, column=c).value or '').lower()
        v_3 = str(sheet.cell(row=6, column=c).value or '').lower()
        v_4 = str(sheet.cell(row=7, column=c).value or '').lower()
        
        v = v_1 + v_2 + v_3 + v_4
        
        if 'tanggal' in v or 'tgl' in v:
            if not col_tgl: col_tgl = c
        if 'mati' in v or 'deplesi' in v or 'mendak' in v:
            if not col_mati: col_mati = c
        if 'afkir' in v or 'jual' in v:
            if not col_afkir: col_afkir = c
            
    print(f'TGL: {col_tgl}, MATI: {col_mati}, AFKIR: {col_afkir}')
    total_deplesi = 0
    base_date = datetime.date(2025, 3, 24)
    ceiling_date = datetime.date(2026, 4, 11)
    
    for r in range(5, 100): 
        t = sheet.cell(row=r, column=col_tgl).value
        # also print population column if exists (usually col 14 or 15)
        # We will just print the MATI/AFKIR
        if isinstance(t, datetime.datetime):
            if t.date() > base_date and t.date() <= ceiling_date:
                m = sheet.cell(row=r, column=col_mati).value
                m_val = int(m) if type(m) in (int, float) else 0
                total_deplesi += m_val
                if m_val > 0:
                    print(f'{t.date()} | Mati: {m_val}')
    print('Total extracted deaths:', total_deplesi)
