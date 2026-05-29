import sys, os, io
sys.path.append(os.getcwd())
from tools.google_drive_tool import GoogleDriveTool
import openpyxl

tool = GoogleDriveTool()
fid = '1IaVF9iIXFfREX1V6d6yzPugYhVoSfEwF'  # KD 5
content = tool.download_file(fid)
wb = openpyxl.load_workbook(io.BytesIO(content.getvalue()), data_only=True)

# Pick the Data Harian sheet
harian = next((wb[n] for n in wb.sheetnames if 'HARIAN' in n.upper()), None)
sheet = harian or wb.worksheets[0]
print('Sheet:', sheet.title)

# Print all non-empty header cells in rows 1-15
print('\n--- Header rows ---')
for r in range(1, 16):
    for c in range(1, 25):
        v = sheet.cell(row=r, column=c).value
        if v is not None and str(v).strip():
            print(f'  ({r},{c}) [col {chr(64+c)}]: {repr(v)}')
