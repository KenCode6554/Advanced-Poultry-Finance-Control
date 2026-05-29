import sys, os, io
sys.path.append(os.getcwd())
from tools.google_drive_tool import GoogleDriveTool
tool = GoogleDriveTool()

fid = None
for q in tool.drive_service.files().list(q="name = 'REC KD 5 PL241P JTP Mojogedang .xlsx'", spaces='drive').execute().get('files', []):
    fid = q['id']

if fid:
    print('Downloading', fid)
    content = tool.download_file(fid)
    import openpyxl, datetime
    wb = openpyxl.load_workbook(io.BytesIO(content.getvalue()), data_only=True)
    sheet_name = [s for s in wb.sheetnames if 'harian' in s.lower()]
    sheet = wb[sheet_name[0]] if sheet_name else wb.worksheets[0]
    
    col_tgl = col_mati = col_afkir = None
    for c in range(1, 30):
        v_1 = str(sheet.cell(row=4, column=c).value or '').lower()
        v_2 = str(sheet.cell(row=5, column=c).value or '').lower()
        
        if 'tanggal' in v_1 or 'tanggal' in v_2 or 'tgl' in v_1 or 'tgl' in v_2:
            col_tgl = c
        if 'mati' in v_1 or 'mati' in v_2 or 'deplesi' in v_1 or 'deplesi' in v_2:
            if not col_mati: col_mati = c
        if 'afkir' in v_1 or 'afkir' in v_2 or 'jual' in v_1 or 'jual' in v_2:
            if not col_afkir: col_afkir = c
            
    print(f'TGL: {col_tgl}, MATI: {col_mati}, AFKIR: {col_afkir}')
    total_deplesi = 0
    base_date = datetime.date(2025, 3, 24)
    ceiling_date = datetime.date(2026, 4, 11)
    
    for r in range(5, 50): # up to 50 rows
        t = sheet.cell(row=r, column=col_tgl).value
        if isinstance(t, datetime.datetime):
            if t.date() > base_date and t.date() <= ceiling_date:
                m = sheet.cell(row=r, column=col_mati).value
                a = sheet.cell(row=r, column=col_afkir).value
                m_val = int(m) if type(m) in (int, float) else 0
                a_val = int(a) if type(a) in (int, float) else 0
                total_deplesi += (m_val + a_val)
                print(f'{t.date()} | Mati: {m_val} | Afkir: {a_val} | Cum: {total_deplesi}')
