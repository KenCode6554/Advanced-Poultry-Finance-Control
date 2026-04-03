import io
import os
import sys
import openpyxl
from dotenv import load_dotenv

# Add the tools directory to sys.path
sys.path.append('c:\\Users\\ASUS\\OneDrive\\Documents\\Adcanded Poultry Finance Control\\tools')
sys.path.append('c:\\Users\\ASUS\\OneDrive\\Documents\\Adcanded Poultry Finance Control')

from google_drive_tool import GoogleDriveTool

def find_importrange_sources():
    load_dotenv('c:\\Users\\ASUS\\OneDrive\\Documents\\Adcanded Poultry Finance Control\\.env')
    tool = GoogleDriveTool()
    file_id = '1St1UKPst2EZMpEJK8bT5-ls3bXwus_4n'
    
    content = tool.download_file(file_id)
    wb = openpyxl.load_workbook(io.BytesIO(content.getvalue()), data_only=False)
    
    for sname in ['Data Harian', 'Data Mingguan', 'Data_Out']:
        if sname in wb.sheetnames:
            sheet = wb[sname]
            print(f"Checking {sname} for IMPORTRANGE:")
            found = False
            for r in range(1, 50):
                for c in range(1, 20):
                    v = str(sheet.cell(row=r, column=c).value)
                    if v and "IMPORTRANGE" in v.upper():
                        print(f"  FOUND at {sname}!{openpyxl.utils.get_column_letter(c)}{r}: {v}")
                        found = True
            if not found:
                print(f"  No IMPORTRANGE found in first 50x20 of {sname}")

if __name__ == "__main__":
    find_importrange_sources()
