
import openpyxl
import io
from google_drive_tool import GoogleDriveTool

tool = GoogleDriveTool()
file_name = "Rec P. fajar kd 6A BBK  (AL 1001).xlsx"
file_id = "1O3p_C5R39S9kNoRIdPz80WvOQpYreY6L" # I need to find the ID or search by name

# Search for the file id
results = tool.drive_service.files().list(q=f"name='{file_name}' and trashed=false", fields="files(id, name)").execute()
if not results.get('files'):
    print("File not found")
    exit()

file_id = results['files'][0]['id']
content = tool.download_file(file_id)
with open("temp_debug.xlsx", "wb") as f:
    f.write(content.getvalue())
wb = openpyxl.load_workbook("temp_debug.xlsx", data_only=False)

for sname in wb.sheetnames:
    ws = wb[sname]
    for r in ws.iter_rows(max_row=100, max_col=30):
        for c in r:
            if c.value and 'IMPORTRANGE' in str(c.value):
                print(f"Sheet: {sname}, Cell: {c.coordinate}, Value: {c.value}")
