"""
debug_jtp_tail.py
Inspect the tail of JTP files to see why activity detection stopped.
"""
import os
import openpyxl
from google_drive_tool import GoogleDriveTool

tool = GoogleDriveTool()
files = [
    {'name': 'REC KD 5 PL241P JTP Mojogedang .xlsx', 'id': '1fX9pX6z6f9n9zX_zX9zX9zX9zX9zX9zX'}, # Need real IDs or just use local search
]

# Since I don't have the IDs handy, I'll search by name in the root
root_id = os.getenv("GOOGLE_DRIVE_ROOT_ID")
folders = tool.get_farm_folders(root_id)
jtp_folder = next(f for f in folders if 'JTP' in f['name'].upper())
jtp_files = tool.list_xlsx_files(jtp_folder['id'])

target_files = ['REC KD 5 PL241P JTP Mojogedang .xlsx', 'REC KD 7 PL241P JTP Mojogedang.xlsx']

for file in jtp_files:
    if file['name'] in target_files:
        print(f"\n--- Investigating {file['name']} ---")
        content_io = tool.download_file(file['id'])
        with open('temp_debug.xlsx', 'wb') as f:
            f.write(content_io.getbuffer())
        
        wb = openpyxl.load_workbook('temp_debug.xlsx', data_only=True, read_only=True)
        sheet_name = next((s for s in wb.sheetnames if 'Harian' in s), wb.sheetnames[0])
        sheet = wb[sheet_name]
        print(f"Sheet used: {sheet_name}")
        
        # Print Headers (Rows 7-11)
        print("Headers:")
        for r in range(7, 12):
            row_data = [sheet.cell(row=r, column=c).value for c in range(1, 30)]
            print(f"  Row {r:2}: {row_data}")

        # Check rows 395 to 417
        print("\nData:")
        for r in range(395, 417):
            row_vals = []
            for c in range(1, 30):
                val = sheet.cell(row=r, column=c).value
                row_vals.append(val)
            print(f"Row {r:3}: {row_vals}")
