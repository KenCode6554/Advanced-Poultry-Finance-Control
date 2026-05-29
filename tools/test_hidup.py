import sys, os, io, datetime
sys.path.append(os.getcwd())
from tools.google_drive_tool import GoogleDriveTool
import openpyxl
tool = GoogleDriveTool()
fid = '1_AVPCyEuVLUulSeUzPYehCqkpZQFrRw0'
content = tool.download_file(fid)
wb = openpyxl.load_workbook(io.BytesIO(content.getvalue()), data_only=True)
sheet = [s for s in wb.worksheets if 'harian' in s.title.lower()][0]

col_tgl = 1   # Tanggal at (9,1)
col_hidup = 5  # Hidup at (9,5)
ceiling_date = datetime.date(2026, 4, 11)

best_pop = None
best_date = None

for r in range(10, sheet.max_row + 1):
    d_val = sheet.cell(row=r, column=col_tgl).value
    row_dt = None
    if isinstance(d_val, datetime.datetime):
        row_dt = d_val.date()
    
    if row_dt and row_dt <= ceiling_date:
        h_val = sheet.cell(row=r, column=col_hidup).value
        if h_val is not None:
            try:
                pop_int = int(float(str(h_val).replace(",", "").strip()))
                if pop_int > 0:
                    best_pop = pop_int
                    best_date = row_dt
            except: pass

print(f'Best Hidup: {best_pop} on {best_date}')
