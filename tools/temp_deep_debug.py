import os
import openpyxl
from google_drive_tool import GoogleDriveTool
import dotenv

dotenv.load_dotenv()

def debug_specifics():
    tool = GoogleDriveTool()
    # IDs from previous logs
    targets = {
        "KD 1A": "1M3TRIdgN9yIynHRdORe3cQUqXsUbWBc3",
        "KD 5": "1v4UqZ-5O8M6m1N2D_9KJH9-H_U-W0X7S", # Just a guess, need to find it
        "KD 6A": "1-U_U-V_U-W_U-X_U-Y", # Need to find
        "KD 17": "1-Z_U-A_U-B" # Need to find
    }
    
    # Let's just list all files in the root and find them
    root_folder = os.getenv("GOOGLE_DRIVE_ROOT_ID")
    farms = tool.get_farm_folders(root_folder)
    
    to_inspect = []
    for farm in farms:
        files = tool.list_xlsx_files(farm['id'])
        for f in files:
            name = f['name'].upper()
            if "KD 5" in name or "KD 1A" in name or "KD 6A" in name or "KD 17" in name or "KD 7" in name:
                to_inspect.append(f)
                
    for f in to_inspect:
        print(f"\n=== INSPECTING {f['name']} ({f['id']}) ===")
        content = tool.download_file(f['id'])
        wb = openpyxl.load_workbook(content, data_only=False)
        target_sheet = None
        for name in wb.sheetnames:
            if 'Harian' in name:
                target_sheet = name
                break
        
        if not target_sheet:
            print("No Data Harian sheet found.")
            continue
            
        ws = wb[target_sheet]
        
        # Look for headers to find columns
        header_row = 7
        cols = {}
        for c in range(1, 15):
            val = ws.cell(row=header_row, column=c).value
            if val:
                cols[str(val).strip()] = c
        print(f"Headers found: {cols}")
        
        # Find the last few rows with formulas or #N/A
        max_r = ws.max_row
        print(f"Max row: {max_r}")
        
        # Dump the last 20 rows of Column F (or whichever has population)
        # Usually Col 6 (F)
        pop_col = 6
        for r in range(max(1, max_r - 20), max_r + 1):
            val_form = ws.cell(row=r, column=pop_col).value
            # Also check date in col 1
            date_val = ws.cell(row=r, column=1).value
            print(f"Row {r} | Date: {date_val} | Pop Formula/Value: {val_form}")
            
            # If we don't see IMPORTRANGE in pop_col, check OTHER columns in this row
            if r > 10 and not (isinstance(val_form, str) and 'IMPORTRANGE' in val_form):
                for c in range(1, 20):
                    other_val = ws.cell(row=r, column=c).value
                    if isinstance(other_val, str) and 'IMPORTRANGE' in other_val:
                        print(f"  !! Found IMPORTRANGE in Col {c}: {other_val}")

if __name__ == "__main__":
    debug_specifics()
