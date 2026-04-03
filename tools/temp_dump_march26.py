import os
import openpyxl
from google_drive_tool import GoogleDriveTool
import dotenv
import datetime

dotenv.load_dotenv()

def debug():
    tool = GoogleDriveTool()
    root_folder = os.getenv("GOOGLE_DRIVE_ROOT_ID")
        
    farms = tool.get_farm_folders(root_folder)
    
    for farm in farms:
        files = tool.list_xlsx_files(farm['id'])
        for file in files:
            name_upper = file['name'].upper()
            if "1A" in name_upper and "BBK" in name_upper:
                print(f"Found {file['name']}, downloading...")
                content = tool.download_file(file['id'])
                wb = openpyxl.load_workbook(content, data_only=True, read_only=True)
                
                for sheet_name in wb.sheetnames:
                    if "Harian" in sheet_name:
                        ws = wb[sheet_name]
                        print(f"Searching in '{sheet_name}'...")
                        
                        all_rows = list(ws.iter_rows(values_only=True))
                        
                        print("\n--- Rows around March 2026 ---")
                        for r_idx, row in enumerate(all_rows):
                            # The date is usually in col 0
                            if len(row) > 0 and isinstance(row[0], datetime.datetime):
                                dt = row[0]
                                if dt.year == 2026 and dt.month == 3 and 10 <= dt.day <= 20:
                                    print(f"Row {r_idx+1} ({dt.strftime('%Y-%m-%d')}): {row[:10]}")

                wb.close()
                return

if __name__ == "__main__":
    debug()
