import os
import openpyxl
from google_drive_tool import GoogleDriveTool
import dotenv
import datetime

dotenv.load_dotenv()

def debug():
    tool = GoogleDriveTool()
    root_folder = os.getenv("GOOGLE_DRIVE_ROOT_ID")
        
    farms = tool.get_farm_folders(root_folder)
    
    for farm in farms:
        files = tool.list_xlsx_files(farm['id'])
        for file in files:
            name_upper = file['name'].upper()
            if "1A" in name_upper and "BBK" in name_upper:
                print(f"Found {file['name']}, downloading...")
                content = tool.download_file(file['id'])
                # READ WITH DATA_ONLY=FALSE to get formulas!
                wb = openpyxl.load_workbook(content, data_only=False, read_only=True)
                
                for sheet_name in wb.sheetnames:
                    if "Harian" in sheet_name:
                        ws = wb[sheet_name]
                        print(f"Searching formulas in '{sheet_name}'...")
                        
                        all_rows = list(ws.iter_rows(values_only=True))
                        
                        print("\n--- Formulas around March 11-16 2026 ---")
                        count = 0
                        for r_idx, row in enumerate(all_rows):
                            # Col 0 might be the date
                            if len(row) > 0:
                                val = row[0]
                                if type(val) is datetime.datetime and val.year == 2026 and val.month == 3 and 10 <= val.day <= 16:
                                    print(f"Row {r_idx+1} ({val.strftime('%Y-%m-%d')}):")
                                    print(f"  Col 2 (Deplesi): {row[2]}")
                                    print(f"  Col 4 (Hidup):   {row[4]}")
                                    print(f"  Col 5 (Produksi Telur): {row[5]}")
                                    count += 1
                        
                        if count == 0:
                            print("Did not find matching dates in Col 0.")
                            # Print row 275-285 just in case
                            for r_idx in range(275, 285):
                                print(f"Row {r_idx}: {all_rows[r_idx][:6]}")

                wb.close()
                return

if __name__ == "__main__":
    debug()
