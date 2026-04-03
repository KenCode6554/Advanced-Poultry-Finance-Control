import os
import openpyxl
from google_drive_tool import GoogleDriveTool
import dotenv

dotenv.load_dotenv()

def debug():
    tool = GoogleDriveTool()
    root_folder = os.getenv("GOOGLE_DRIVE_ROOT_ID")
    if not root_folder:
        print("No GOOGLE_DRIVE_ROOT_ID")
        return
        
    farms = tool.get_farm_folders(root_folder)
    
    for farm in farms:
        files = tool.list_xlsx_files(farm['id'])
        for file in files:
            name_upper = file['name'].upper()
            if "KD 4" in name_upper and "JTP" in name_upper:
                print(f"Found {file['name']}, downloading...")
                content = tool.download_file(file['id'])
                wb = openpyxl.load_workbook(content, data_only=True, read_only=True)
                
                for sheet_name in wb.sheetnames:
                    if "Harian" in sheet_name:
                        ws = wb[sheet_name]
                        print(f"Searching in '{sheet_name}'...")
                        
                        all_rows = list(ws.iter_rows(values_only=True))
                        
                        # Find headers
                        print("\n--- Rows 7 to 12 (Headers potential) ---")
                        for r_idx in range(6, 12):
                            if r_idx < len(all_rows):
                                print(f"Row {r_idx+1}: {all_rows[r_idx][:10]}")
                                
                        # Find value 3342
                        print("\n--- Hunting for 3342 ---")
                        for r_idx, row in enumerate(all_rows):
                            for c_idx, val in enumerate(row):
                                if val == 3342 or str(val).strip() == '3342' or val == 3342.0:
                                    print(f"FOUND 3342 at Row {r_idx+1}, Col {c_idx} (0-indexed)")
                                    print(f"  Row {r_idx+1} full: {row[:10]}")
                        
                        print("\n--- Last 5 rows ---")
                        for r_idx in range(len(all_rows)-5, len(all_rows)):
                            if r_idx >= 0:
                                print(f"Row {r_idx+1}: {all_rows[r_idx][:10]}")

                wb.close()
                return

if __name__ == "__main__":
    debug()
