import sys, os, io
sys.path.append(os.getcwd())
from tools.google_drive_tool import GoogleDriveTool
import openpyxl

tool = GoogleDriveTool()
fid = '1_AVPCyEuVLUulSeUzPYehCqkpZQFrRw0'
content = tool.download_file(fid)
wb = openpyxl.load_workbook(io.BytesIO(content.getvalue()), data_only=True)
sheet = next((wb[n] for n in wb.sheetnames if any(x in n.upper() for x in ["DATA", "HARIAN"])), wb.active)
print('Sheet found:', sheet.title)

col_tgl = None
col_hidup = None
header_row_idx = 9

for r in range(1, 16):
    row_vals = [str(sheet.cell(row=r, column=c).value or '').upper().strip() for c in range(1, 30)]
    has_tgl = 'TANGGAL' in row_vals or 'TGL' in row_vals
    print(f'Row {r}: has_TANGGAL={has_tgl}', [v for v in row_vals if v])
    if has_tgl:
        header_row_idx = r
        for c_idx, v in enumerate(row_vals):
            if v in ('TANGGAL', 'TGL') and col_tgl is None:
                col_tgl = c_idx + 1
                print(f'  -> col_tgl = {col_tgl}')
        for sr in range(max(1, r - 2), r + 3):
            for c in range(1, 30):
                v = str(sheet.cell(row=sr, column=c).value or '').upper().strip()
                if 'HIDUP' in v:
                    print(f'  -> HIDUP found at ({sr},{c}) = {repr(sheet.cell(row=sr, column=c).value)}')
                    if col_hidup is None:
                        col_hidup = c
        break

print(f'\nFinal: col_tgl={col_tgl}, col_hidup={col_hidup}')
