import os
import json
import io
import re
import pandas as pd
import openpyxl
import sys
from datetime import datetime, date, timedelta
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from dotenv import load_dotenv

load_dotenv()

class GoogleDriveTool:
    def __init__(self):
        self.creds_path = os.getenv('GOOGLE_APPLICATION_CREDENTIALS')
        self.creds_json = os.getenv('GOOGLE_SERVICE_ACCOUNT_JSON')
        self.creds = None
        self.drive_service = None
        self.sheets_service = None
        
        self.MIME_SHEETS = 'application/vnd.google-apps.spreadsheet'
        self.MIME_XLSX = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        # Priority 1: GOOGLE_APPLICATION_CREDENTIALS (file path)
        if self.creds_path and os.path.exists(self.creds_path):
            print(f"   [DRIVE] Loading service account from file: {self.creds_path}")
            self.creds = service_account.Credentials.from_service_account_file(
                self.creds_path, scopes=[
                    'https://www.googleapis.com/auth/drive',
                    'https://www.googleapis.com/auth/spreadsheets.readonly'
                ]
            )
        # Priority 2: GOOGLE_SERVICE_ACCOUNT_JSON (raw JSON string)
        elif self.creds_json:
            print(f"   [DRIVE] Loading service account from environment variable string...")
            try:
                info = json.loads(self.creds_json)
                self.creds = service_account.Credentials.from_service_account_info(
                    info, scopes=[
                        'https://www.googleapis.com/auth/drive',
                        'https://www.googleapis.com/auth/spreadsheets.readonly'
                    ]
                )
            except Exception as e:
                print(f"   [DRIVE] Error parsing GOOGLE_SERVICE_ACCOUNT_JSON: {e}")

        if self.creds:
            print("   [DRIVE] Initializing Google Drive service (Discovery v3)...")
            sys.stdout.flush()
            self.drive_service = build('drive', 'v3', credentials=self.creds)
            
            print("   [DRIVE] Initializing Google Sheets service (Discovery v4)...")
            sys.stdout.flush()
            self.sheets_service = build('sheets', 'v4', credentials=self.creds)
            print("   [DRIVE] Google Services initialized successfully.")
        else:
            print(f"WARNING: No Google credentials found (check GOOGLE_APPLICATION_CREDENTIALS or GOOGLE_SERVICE_ACCOUNT_JSON)")

    def get_farm_folders(self, root_folder_id):
        query = f"'{root_folder_id}' in parents and mimeType = 'application/vnd.google-apps.folder' and trashed = false"
        results = self.drive_service.files().list(q=query, fields="files(id, name)").execute()
        return results.get('files', [])

    def list_google_sheets(self, folder_id):
        """Finds native Google Sheets files directly inside the given folder_id."""
        print(f"Searching for Google Sheets in folder {folder_id}...")
        query = f"'{folder_id}' in parents and mimeType = '{self.MIME_SHEETS}' and trashed = false"
        return self._list_files(query, folder_id)

    def list_binary_xlsx_files(self, folder_id):
        """Finds binary .xlsx files that likely need to be converted to Google Sheets."""
        query = f"'{folder_id}' in parents and mimeType = '{self.MIME_XLSX}' and trashed = false"
        return self._list_files(query, folder_id)

    def _list_files(self, query, folder_id):
        results = []
        page_token = None
        while True:
            response = self.drive_service.files().list(
                q=query, 
                fields="nextPageToken, files(id, name, parents, mimeType, modifiedTime)",
                pageToken=page_token
            ).execute()
            
            items = response.get('files', [])
            for f in items:
                name = f['name'].upper()
                if any(x in name for x in ["BBK", "JTP", "KD", "REC"]):
                    results.append(f)
            
            page_token = response.get('nextPageToken')
            if not page_token:
                break
        return results

    def list_xlsx_files(self, folder_id):
        """DEPRECATED: Use list_google_sheets instead for live sync."""
        return self.list_binary_xlsx_files(folder_id)

    def download_file(self, file_id):
        # Determine if it's a native Google Sheet to use Export instead of GetMedia
        file_metadata = self.drive_service.files().get(fileId=file_id, fields='mimeType').execute()
        mime_type = file_metadata.get('mimeType')
        
        if mime_type == self.MIME_SHEETS:
            print(f"      [DRIVE] Exporting Google Sheet to .xlsx for processing...")
            request = self.drive_service.files().export_media(
                fileId=file_id,
                mimeType=self.MIME_XLSX
            )
        else:
            request = self.drive_service.files().get_media(fileId=file_id)
            
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
        fh.seek(0)
        return fh

    def _safe_float(self, val):
        if pd.isna(val) or str(val).strip() == '': return 0.0
        s = str(val).strip().upper()
        # Handle common Excel error strings and Loading indicator
        if any(x in s for x in ['#DIV/0!', '#REF!', '#N/A', '#VALUE!', 'LOADING...', 'NAN', 'NONE', 'NULL']):
            return 0.0
        try:
            # Remove % and , but keep .
            s = s.replace(',', '').replace('%', '').strip()
            return float(s)
        except: 
            return 0.0

    def _clamp(self, val, precision, scale):
        if val is None: return None
        limit = 10**(precision-scale) - (1/10**scale)
        try:
            f = float(val)
            if pd.isna(f): return None
            return max(-limit, min(limit, f))
        except: return None

    def _find_column_indices(self, df, kandang_name="Unknown"):
        """Dynamically identify column indices for various metrics from the first 150 rows."""
        indices = {}

        # We'll use multiple passes over the first 150 rows to build indices
        # Pass 1: Primary headers (Row 7-9 usually)
        for i in range(min(12, len(df))):
            # Skip the very first rows that usually contain metadata/titles to avoid collisions (e.g. "2 UMUR" matching week)
            if i < 7: continue 
            
            row_vals = [str(x).upper().strip() for x in df.iloc[i].values]
            prev_row = [str(x).upper().strip() for x in df.iloc[i-1].values] if i > 0 else []
            
            def find_col(keywords, priority_keywords=None):
                if priority_keywords:
                    for kw in priority_keywords:
                        if kw in row_vals: return row_vals.index(kw)
                for kw in keywords:
                    for idx, val in enumerate(row_vals):
                        if kw in val: return idx
                return None

            # Date (week end date)
            d_idx = find_col(['TANGGAL', 'TGL', 'AKHIR MG', 'AKHIRMG', 'DATE'], ['TANGGAL', 'AKHIR MG'])
            if d_idx is not None and 'date' not in indices: indices['date'] = d_idx

            # Week
            w_idx = find_col(['AKHIR MG', 'WEEK', 'MGG', 'USIA', 'UMUR'], ['AKHIR MG', 'USIA (MG)', 'UMUR'])
            if w_idx is not None and 'week' not in indices: indices['week'] = w_idx
            
            # HD%
            hd_idx = find_col(['% HD', 'HD %', 'HD%', 'ACTUAL %', 'HD'], ['% HD', 'ACTUAL %'])
            if hd_idx is not None and 'hd_act' not in indices: indices['hd_act'] = hd_idx
            
            # Egg Weight (Grams/Butir — g/btr column ONLY)
            # IMPORTANT: do NOT match 'Kg' (total kg) or 'Kg/1000' (egg mass).
            # Only look for explicit 'G/BTR' or 'g/btr' which is grams-per-egg.
            ew_idx = None
            for ci, val in enumerate(row_vals):
                # Exact match only — 'G/BTR' must be the cell value (not a substring of Kg, Kum, etc.)
                if val in ['G/BTR', 'G BTR', 'GRAM/BTR', 'EGG WEIGHT (G)', 'BERAT BTR']:
                    ew_idx = ci
                    break
            if ew_idx is not None and 'ew_act' not in indices:
                indices['ew_act'] = ew_idx
                if ew_idx + 1 < len(row_vals) and 'STD' in row_vals[ew_idx + 1]:
                    indices['ew_std'] = ew_idx + 1

            # Berat Telur (Kg/1000) — first pass: look in this row
            em_idx = None
            if 'KG/1000' in row_vals: em_idx = row_vals.index('KG/1000')
            elif 'BE/1000' in row_vals: em_idx = row_vals.index('BE/1000')
            else:
                for idx, val in enumerate(row_vals):
                    if 'KG/1000' in val and 'EK' not in val:
                        em_idx = idx; break
            
            if em_idx is not None and 'em_act' not in indices:
                indices['em_act'] = em_idx
                if em_idx + 1 < len(row_vals) and 'STD' in row_vals[em_idx + 1]:
                    indices['em_std'] = em_idx + 1

            # Berat Telur parent-header scan: find 'BERAT TELUR' group → look for Kg/1000 (no Ekor) sub-cols
            # This overrides the first-pass result when the spreadsheet has a multi-row 'Berat Telur' header.
            # Layout (all templates): Row 8 = 'Berat Telur', Row 9: Kg | Kum | Cumm | g/btr | Std | Cum | Kg/1000 ek | Act | Kg/1000 | Std
            # We want the 'Kg/1000' col (NO 'ek' suffix and NO 'EKOR' in the row below).
            berat_telur_col = None
            for ci, v in enumerate(row_vals):
                if 'BERAT TELUR' in v or ('BERAT' in v and ci < len(prev_row) and 'TELUR' in prev_row[ci]):
                    berat_telur_col = ci; break
            if berat_telur_col is None:  # also check prev_row (parent header one row above)
                for ci, pv in enumerate(prev_row):
                    if 'BERAT TELUR' in pv:
                        berat_telur_col = ci; break
            if berat_telur_col is not None:
                next_row_vals = [str(x).upper().strip() for x in df.iloc[i+1].values] if i+1 < len(df) else []
                for c in range(berat_telur_col, min(berat_telur_col + 15, len(row_vals))):
                    rv = row_vals[c]
                    if 'KG/1000' in rv:
                        # Skip if the cell itself contains 'EK' (= 'Kg/1000 ek' → Egg Mass, not Berat Telur)
                        if 'EK' in rv: continue
                        # Skip if the row below this cell has 'EKOR' (same reason)
                        below = next_row_vals[c] if c < len(next_row_vals) else ''
                        if 'EKOR' in below: continue
                        # This is the true Berat Telur Kg/1000
                        indices['em_act'] = c
                        if c + 1 < len(row_vals) and 'STD' in row_vals[c + 1]:
                            indices['em_std'] = c + 1
                        break

            # Feed Intake (G/E/H or FC)
            geh_idx = None
            for idx, val in enumerate(row_vals):
                # Look for EXACT matches or specific keywords, skipping FCR
                if val in ['G/E/H', 'G/E/HR', 'GRAM/EKOR', 'FC', 'FK'] and 'FCR' not in val:
                    geh_idx = idx; break
                # Fallback to substring only if not FCR
                if any(k in val for k in ['G/E/H', 'G/E/HR']) and 'FCR' not in val:
                    geh_idx = idx; break
            
            if geh_idx is not None and 'pakan_act' not in indices:
                # Ensure it's not a standard column itself 
                if 'STD' not in row_vals[geh_idx]:
                    indices['pakan_act'] = geh_idx
                    # Find standard (look left or right)
                    for offset in [1, -1]:
                        target = geh_idx + offset
                        if 0 <= target < len(row_vals) and ('STD' in row_vals[target] or 'STAND' in row_vals[target]):
                            indices['pakan_std'] = target; break
                elif 'pakan_std' not in indices:
                    indices['pakan_std'] = geh_idx

            # Pakan KG
            pk_kg_idx = find_col(['PAKAN KG', 'KONS. KG', 'TOTAL KG'], ['KG'])
            if pk_kg_idx is not None and 'pakan_kg' not in indices:
                if '1000' not in row_vals[pk_kg_idx] and 'BUTIR' not in row_vals[pk_kg_idx] and 'FCR' not in row_vals[pk_kg_idx]:
                    indices['pakan_kg'] = pk_kg_idx

            # FCR
            fcr_parent = None
            for idx, val in enumerate(row_vals):
                if idx > 35: continue # Primary metrics are always in the first 35 columns
                if 'FCR' in val or (idx < len(prev_row) and 'FCR' in prev_row[idx]):
                    if 'ACT' in val or 'ACTUAL' in val or (idx < len(prev_row) and 'ACT' in prev_row[idx]):
                         if 'fcr_act' not in indices: indices['fcr_act'] = idx
                    elif 'CUM' in val or 'CUMM' in val or (idx < len(prev_row) and 'CUM' in prev_row[idx]):
                         if 'fcr_cum' not in indices: indices['fcr_cum'] = idx

            # Relative search for FCR Cumulative if Actual was found but Cum was not
            if 'fcr_act' in indices and 'fcr_cum' not in indices:
                fa_idx = indices['fcr_act']
                # Search immediate right (up to 2 columns) for CUM/CUMM or even NAN if it's implicitly there
                for offset in [1, 2]:
                    if fa_idx + offset < len(row_vals):
                        v = row_vals[fa_idx + offset]
                        pv = prev_row[fa_idx + offset] if fa_idx + offset < len(prev_row) else ""
                        if any(k in v for k in ['CUM', 'KUM']) or any(k in pv for k in ['CUM', 'KUM']):
                            indices['fcr_cum'] = fa_idx + offset
                            break
            
            # If still missing FCR, try specific keywords (Restrict to production area < 35)
            if 'fcr_act' not in indices:
                fa_idx = find_col(['FCR ACTUAL', 'FCR ACT'], None)
                if fa_idx is not None and fa_idx < 35: indices['fcr_act'] = fa_idx
            if 'fcr_cum' not in indices:
                # Use strict keyword first for FCR specific columns
                fc_idx = find_col(['FCR CUMM', 'FCR CUM'], None)
                if fc_idx is not None and fc_idx < 35: indices['fcr_cum'] = fc_idx

            # Deplesi — Strategy 1: direct keyword in this row
            dep_pct = find_col(['% CUM', 'CUM %', 'DEPLESI %'], ['% CUM'])
            if dep_pct is not None and 'deplesi_pct' not in indices: indices['deplesi_pct'] = dep_pct
            dep_ekor = find_col(['DEP EKOR', 'DEPLESI EKOR', 'MATI EKOR'], [])
            if dep_ekor is not None and 'deplesi_ekor' not in indices: indices['deplesi_ekor'] = dep_ekor

        # Deplesi — Strategy 2: find 'DEPLESI' as a parent header and scan sub-rows beneath it
        if 'deplesi_ekor' not in indices or 'deplesi_pct' not in indices:
            for i in range(min(12, len(df))):
                row_vals_raw = [str(x).upper().strip() for x in df.iloc[i].values]
                for col_i, val in enumerate(row_vals_raw):
                    if val == 'DEPLESI':
                        # Scan up to 2 rows below this header for sub-column labels
                        weekly_pct_col = None  # fallback
                        for sub_row in range(i+1, min(i+3, len(df))):
                            sub_vals = [str(x).upper().strip() for x in df.iloc[sub_row].values]
                            # also get the next sub-row for two-row header detection
                            next_sub_vals = [str(x).upper().strip() for x in df.iloc[sub_row+1].values] if sub_row+1 < len(df) else []
                            for c in range(col_i, min(col_i + 8, len(sub_vals))):
                                sv = sub_vals[c]
                                if 'EKOR' in sv and 'deplesi_ekor' not in indices:
                                    indices['deplesi_ekor'] = c
                                # Cumulative ekor count: a 'CUM'/'KUM' cell that is NOT a % cell
                                if sv.strip() in ['CUM', 'KUM'] and 'deplesi_cum' not in indices:
                                    # Make sure the next sub-row for this col doesn't have '%' — that would be %Cum not Cum ekor
                                    below = next_sub_vals[c] if c < len(next_sub_vals) else ''
                                    if '%' not in below:
                                        indices['deplesi_cum'] = c
                                # Cumulative %: a ' CUM'/' KUM' cell whose next sub-row has '%'
                                if ('CUM' in sv or 'KUM' in sv) and 'deplesi_pct' not in indices:
                                    below = next_sub_vals[c] if c < len(next_sub_vals) else ''
                                    if '%' in below:
                                        indices['deplesi_pct'] = c
                                # Remember first standalone '%' as fallback for weekly pct
                                if sv == '%' and weekly_pct_col is None:
                                    weekly_pct_col = c
                        # Only use weekly % as fallback if cumulative % was NOT found
                        if 'deplesi_pct' not in indices and weekly_pct_col is not None:
                            indices['deplesi_pct'] = weekly_pct_col
                        break

        # Refined Pass for Deplesi Count if pct found (original fallback)
        if 'deplesi_pct' in indices and 'deplesi_cum' not in indices:
            dep_pct_col = indices['deplesi_pct']
            # Look left for cumulative bird count
            for j in range(dep_pct_col - 1, max(-1, dep_pct_col - 4), -1):
                col_sample = [str(x).upper() for x in df.iloc[7:10, j].values if pd.notna(x)]
                if any('CUM' in s or 'EKOR' in s for s in col_sample):
                    indices['deplesi_cum'] = j; break
        if 'deplesi_pct' in indices and 'deplesi_ekor' not in indices:
            dep_pct_col = indices['deplesi_pct']
            for j in range(dep_pct_col - 1, max(-1, dep_pct_col - 5), -1):
                col_sample = [str(x).upper() for x in df.iloc[7:10, j].values if pd.notna(x)]
                if 'EKOR' in ' '.join(col_sample) and 'CUM' not in ' '.join(col_sample):
                    indices['deplesi_ekor'] = j; break

        # Fill in missing standards by looking adjacent to Actuals if not found
        for key in ['ew_act', 'em_act', 'pakan_act', 'fcr_act', 'hd_act']:
            std_key = key.replace('_act', '_std') if '_act' in key else key + '_std'
            if key in indices and std_key not in indices:
                idx = indices[key]
                # Look adjacent (1-3 cols) for a column containing 'STD'
                for j in range(idx + 1, min(idx + 4, len(df.columns))):
                    col_sample = [str(x).upper() for x in df.iloc[:10, j].values if pd.notna(x)]
                    if any('STD' in s or 'STAND' in s or 'STND' in s for s in col_sample):
                        indices[std_key] = j; break
        return indices

    def _safe_float(self, val):
        if val is None: return None
        s = str(val).strip().upper()
        if any(x in s for x in ["LOADING", "#N/A", "#VALUE!", "#REF!", "#DIV/0!", "NAN"]): 
            return None
        try:
            if isinstance(val, str):
                val = val.replace(',', '').replace('%', '').strip()
            if not val:
                return None
            f_val = float(val)
            return f_val if pd.notna(f_val) else None
        except:
            return None

    def _read_dataout_via_sheets_api(self, file_id):
        """Reads Data_Out sheet live via the Google Sheets API to bypass stale xlsx cached formula values."""
        try:
            # Get the spreadsheet metadata to find the Data_Out sheet name
            meta = self.sheets_service.spreadsheets().get(spreadsheetId=file_id).execute()
            sheet_name = next(
                (s['properties']['title'] for s in meta.get('sheets', [])
                 if 'DATA_OUT' in s['properties']['title'].upper() or 'DATA OUT' in s['properties']['title'].upper()),
                None
            )
            if not sheet_name:
                return None
            resp = self.sheets_service.spreadsheets().values().get(
                spreadsheetId=file_id,
                range=f"'{sheet_name}'!A1:AZ500",
                valueRenderOption='UNFORMATTED_VALUE',
                dateTimeRenderOption='FORMATTED_STRING'
            ).execute()
            rows = resp.get('values', [])
            if not rows:
                return None
            # Pad rows to equal length and build a DataFrame
            if not rows: return None
            df = pd.DataFrame(rows)
            idx = self._find_harian_columns(df)
            if 'date' not in idx or 'btr' not in idx: return None
            daily_data = []
            for i in range(1, len(df)):
                row = df.iloc[i]
                try:
                    dt = pd.to_datetime(row[idx['date']])
                    btr = float(row[idx['btr']])
                    if pd.notna(dt) and populasi and populasi > 0:
                        hd = (btr / populasi) * 100
                        daily_data.append({'date': dt.strftime('%Y-%m-%d'), 'hd': round(hd, 2)})
                except: continue
            return daily_data
        except Exception as e:
            print(f"  [Sheets API] Failed to read Daily Harian for {file_id}: {e}")
            return None

    def extract_data_from_excel(self, file_content, farm_name, file_name, file_id=None, populasi=None):
        """Extracts production data from an Excel file stream."""
        kandang_name = re.sub(r'^(REC\s+KD\s+|Rec\s+P\.\s+fajar\s+kd\s+|Recording\s+)', '', file_name, flags=re.IGNORECASE)
        kandang_name = re.sub(r'^(REC\s+KD\s+|Rec\s+P.\s+fajar\s+kd\s+|Recording\s+)', '', file_name, flags=re.IGNORECASE)
        kandang_name = kandang_name.replace('.xlsx', '').strip()
        data = {'farm': farm_name, 'kandang': kandang_name, 'weekly': [], 'daily': [], 'populasi': populasi or 0}
        try:
            df = None
            if file_id and self.sheets_service:
                print(f"  [Sheets API] Fetching Data_Out live for {kandang_name}...")
                df = self._read_dataout_via_sheets_api(file_id)
            
            if df is None:
                print(f"  [Sheets API] Live fetch failed or unavailable. Falling back to .xlsx stream for {kandang_name}...")
                df = pd.read_excel(file_content, sheet_name='Data_Out', header=None)

            idx = self._find_column_indices(df, kandang_name)
            
            def get_val(row, key, default=None):
                if key in idx:
                    try: return self._safe_float(row[idx[key]])
                    except: return default
                return default

            if 'week' in idx and 'hd_act' in idx:
                # First pass: find max date in the sheet to handle historical files
                max_sheet_date = datetime.now().date()
                sheet_dates = []
                for i in range(7, min(500, len(df))):
                    try:
                        d = pd.to_datetime(df.iloc[i, 0])
                        if not pd.isna(d): sheet_dates.append(d.date())
                    except: continue
                if sheet_dates:
                    max_sheet_date = max(sheet_dates)

                for i in range(7, len(df)):
                    row = df.iloc[i]
                    if len(row) < 5: continue
                    # Check Column B (index 1) first, then Column C (index 2)
                    # Determine date using idx['date'] if available, fallback to indices 1, 2
                    date_val = row[idx.get('date', 1)] if len(row) > max(idx.get('date', 1), 1) else None
                    if (pd.isna(date_val) or date_val is None) and 'date' not in idx:
                        date_val = row[2] if len(row) > 2 else None
                    
                    # Stop if we hit a secondary header (e.g. "Tanggal" or "Akhir Mg" repeating)
                    row_text = ' '.join([str(x).upper() for x in row.values if pd.notna(x)])
                    if i > 15 and ('TANGGAL' in row_text or 'AKHIR MG' in row_text or 'UMUR (MG)' in row_text):
                        print(f"      [DRIVE] Extraction break at row {i}: Secondary header detected.")
                        break

                    if not isinstance(date_val, (datetime, date)) and (pd.isna(date_val) or date_val == 0): continue
                    try:
                        dt_obj = pd.to_datetime(date_val)
                        date_str = dt_obj.strftime('%Y-%m-%d')
                        # is_future is based on the spreadsheet's own "now" if it's an old file
                        # but we still want to capped at real now for truly future rows
                        is_future = dt_obj.date() > min(max_sheet_date, datetime.now().date())
                    except Exception as e:
                        print(f"      [DEBUG] pd.to_datetime failed for {repr(date_val)}: {e}")
                        continue
                    
                    def safe_int(v):
                        try:
                            if v is None: return 0
                            return int(float(v))
                        except: return 0

                    week_val = row[idx['week']]
                    week_no = 0 if isinstance(week_val, (datetime, date)) else self._safe_float(week_val)
                    if week_no is None or week_no == 0 or week_no > 95: continue
                    
                    def get_act_val(key, c1, c2, threshold=0.0):
                        v = get_val(row, key)
                        if v is not None and v > threshold and not is_future:
                            return self._clamp(v, c1, c2)
                        return None

                    hd_act_val = get_act_val('hd_act', 5, 2)
                    if hd_act_val is not None and 0 < hd_act_val <= 1.0:
                        hd_act_val = self._clamp(hd_act_val * 100, 5, 2)
                    
                    # Harian fallback for HD% — ONLY when the cell contains a formula
                    # placeholder like 'LOADING...' or '#N/A' (i.e. data exists but hasn't
                    # resolved yet). If the raw cell is genuinely NaN/empty, the week hasn't
                    # been recorded yet — skip it rather than pulling partial harian data.
                    if hd_act_val is None and not is_future:
                        raw_hd_cell = row[idx['hd_act']] if 'hd_act' in idx else None
                        cell_str = str(raw_hd_cell).strip().upper() if raw_hd_cell is not None else ''
                        is_placeholder = (
                            cell_str not in ('', 'NAN', 'NONE', '0', '0.0')
                            and not (raw_hd_cell != raw_hd_cell)  # not NaN
                        )
                        if is_placeholder:
                            if 'harian_data' not in locals():
                                file_content.seek(0)
                                harian_data = self._get_harian_data_full(file_content, populasi=populasi)
                            if harian_data and int(week_no) in harian_data:
                                har = harian_data[int(week_no)]
                                hd_val = har.get('hd') if isinstance(har, dict) else har
                                hd_act_val = self._clamp(hd_val, 5, 2)

                    hd_std = get_val(row, 'hd_std')
                    if hd_std is not None and 0 < hd_std <= 1.0:
                        hd_std = self._clamp(hd_std * 100, 5, 2)
                    
                    # Determine if this row has actual production activity
                    has_activity = (hd_act_val is not None) 

                    # Include row if there's actual data OR if there's standard data up to week 90
                    # Exclude future rows entirely — Data_Out pre-fills std values for all 90
                    # weeks, including future ones. Future rows have no actual data and should
                    # never enter the weekly list (they get filtered again in incremental_sync,
                    # but excluding them here saves memory and avoids confusion).
                    if not is_future and (hd_act_val is not None or (week_no <= 90 and (hd_std or 0) > 0)):
                        data['weekly'].append({
                            'date': date_str, 'usia_minggu': int(week_no),
                            'hd_actual': hd_act_val,
                            'hd_std': self._clamp(hd_std, 5, 2),
                            'egg_weight_actual': (lambda v: v if v is not None and 30 <= v <= 100 else None)(get_act_val('ew_act', 6, 2)) if has_activity else None,
                            'egg_weight_std': (lambda v: v if v is not None and 30 <= v <= 100 else None)(self._clamp(get_val(row, 'ew_std'), 6, 2)),
                            'egg_mass_actual': get_act_val('em_act', 6, 2) if has_activity else None,
                            'egg_mass_std': self._clamp(get_val(row, 'em_std'), 6, 2),
                            'fcr_actual': get_act_val('fcr_act', 6, 3) if has_activity else None,
                            'fcr_std': self._clamp(get_val(row, 'fcr_std'), 6, 3),
                            'fcr_cum': self._clamp(get_val(row, 'fcr_cum'), 6, 3),
                            'pakan_kg': get_act_val('pakan_kg', 8, 2) if has_activity else None,
                            'pakan_g_per_ekor_hr': get_act_val('pakan_act', 6, 2) if has_activity else None,
                            'pakan_std': self._clamp(get_val(row, 'pakan_std'), 6, 2),
                            'deplesi_ekor': safe_int(get_val(row, 'deplesi_ekor')) if has_activity else None,
                            'deplesi_cum': safe_int(get_val(row, 'deplesi_cum')) if has_activity else None,
                            'deplesi_pct': self._clamp(get_val(row, 'deplesi_pct'), 5, 2) if has_activity else None
                        })
        except Exception as e:
            print(f"Error extracting {file_name}: {e}")
            import traceback
            traceback.print_exc()
            
        # Append daily data
        if file_id and self.sheets_service:
            daily_rows = self._extract_daily_production_api(file_id, populasi)
            if daily_rows:
                data['daily'] = daily_rows
                
        return data

    def _find_harian_columns(self, df):
        indices = {}
        for r in range(min(15, len(df))):
            row = [str(x).upper() for x in df.iloc[r]]
            for i, val in enumerate(row):
                if any(x in val for x in ['TGL', 'TANGGAL']) and 'date' not in indices: indices['date'] = i
                if any(x in val for x in ['USIA', '(MINGGU)']) and 'week' not in indices and i != indices.get('date'): indices['week'] = i
                if any(x in val for x in ['HD%', '%HD', 'HD (%)']) and 'hd' not in indices: indices['hd'] = i
                if any(x in val for x in ['TOTAL BTR', 'HASIL BTR', 'PRODUKSI TELUR']) and 'btr' not in indices and 'HD%' not in val: indices['btr'] = i
                if any(x in val for x in ['PAKAN KG', 'KG PAKAN']) and 'pakan_kg' not in indices: indices['pakan_kg'] = i
                if any(x in val for x in ['G/EKOR', 'GRAM/EKOR', 'G/E/H']) and 'pakan_g' not in indices: indices['pakan_g'] = i
                # egg weight in grams per butir
                if any(x in val for x in ['G/BTR', 'BERAT BUTIR', 'G BTR']) and 'ew' not in indices: indices['ew'] = i
                # eggs in kg per 1000 (berat telur)
                if any(x in val for x in ['KG/1000', 'KG BTR/1000']) and 'em' not in indices: indices['em'] = i
                # deplesi fields
                if any(x in val for x in ['MATI', 'DEPLESI EKOR', 'DEP EKOR']) and 'dep_ekor' not in indices: indices['dep_ekor'] = i
                if any(x in val for x in ['% DEPLESI', 'DEPLESI %', '% CUM', '%DEPLESI']) and 'dep_pct' not in indices: indices['dep_pct'] = i
                # FCR
                if 'FCR' in val and 'CUM' not in val and 'fcr' not in indices: indices['fcr'] = i
        return indices

    def _get_harian_data_full(self, file_content, populasi=None):
        """Loads Data Harian sheet and returns a mapping of week -> dict of aggregated week metrics.
        Used as fallback when Data_Out shows 'Loading...' formula placeholders."""
        try:
            wb = openpyxl.load_workbook(file_content, data_only=True)
            harian_sheet = next(
                (wb[n] for n in wb.sheetnames if any(x in n.upper() for x in ["HARIAN"])),
                next((wb[n] for n in wb.sheetnames if any(x in n.upper() for x in ["DATA"])), None)
            )
            if not harian_sheet: return None
            
            data = [[cell for cell in row] for row in harian_sheet.iter_rows(values_only=True)]
            df = pd.DataFrame(data)
            
            idx = self._find_harian_columns(df)
            if 'week' not in idx: return None
            
            pop = populasi or 2000  # default population
            
            # Aggregate daily rows -> weekly sums
            week_data = {}
            
            for i in range(len(df)):
                try:
                    w = int(float(str(df.iloc[i, idx['week']])))
                    if w <= 0 or w > 95: continue
                except: continue
                
                if w not in week_data:
                    week_data[w] = {
                        'eggs': 0, 'pakan_kg': 0, 'pakan_g_days': [],
                        'ew_vals': [], 'dep_ekor': 0, 'days': 0, 'fcr_vals': [],
                        'hd_vals': [],  # collect daily HD% values directly from column
                    }
                
                # --- HD% : read directly from the HD% column (most accurate) ---
                if 'hd' in idx:
                    try:
                        hd_raw = self._safe_float(df.iloc[i, idx['hd']])
                        if hd_raw is not None and hd_raw > 0:
                            # normalize: values stored as fraction (0-1) -> convert to pct
                            hd_pct = hd_raw * 100 if hd_raw <= 1.0 else hd_raw
                            week_data[w]['hd_vals'].append(hd_pct)
                    except: pass
                
                # --- egg count (fallback for HD% and for egg_mass) ---
                if 'btr' in idx:
                    try:
                        eggs = float(str(df.iloc[i, idx['btr']] or 0))
                        if eggs > 0:
                            week_data[w]['eggs'] += eggs
                            week_data[w]['days'] += 1
                    except: pass
                
                if 'pakan_kg' in idx:
                    try:
                        pk = self._safe_float(df.iloc[i, idx['pakan_kg']])
                        if pk: week_data[w]['pakan_kg'] += pk
                    except: pass
                
                if 'pakan_g' in idx:
                    try:
                        pg = self._safe_float(df.iloc[i, idx['pakan_g']])
                        if pg: week_data[w]['pakan_g_days'].append(pg)
                    except: pass
                
                if 'ew' in idx:
                    try:
                        ew = self._safe_float(df.iloc[i, idx['ew']])
                        if ew: week_data[w]['ew_vals'].append(ew)
                    except: pass
                
                if 'dep_ekor' in idx:
                    try:
                        de = int(float(str(df.iloc[i, idx['dep_ekor']] or 0)))
                        week_data[w]['dep_ekor'] += de
                    except: pass
                
                if 'fcr' in idx:
                    try:
                        fcr = self._safe_float(df.iloc[i, idx['fcr']])
                        if fcr: week_data[w]['fcr_vals'].append(fcr)
                    except: pass
            
            # Build weekly result dict
            cum_dep = 0
            results = {}
            for w in sorted(week_data.keys()):
                d = week_data[w]
                eggs = d['eggs']
                days = max(d['days'], 1)
                
                # HD%: prefer direct column values; fall back to eggs/pop/days
                if d['hd_vals']:
                    avg_hd = sum(d['hd_vals']) / len(d['hd_vals'])
                    if len(d['hd_vals']) >= 4:
                        # At least 4 days recorded — trust the average
                        hd = avg_hd
                    elif avg_hd >= 20.0:
                        # Fewer days but average is reasonable (e.g. early ramp-up weeks)
                        hd = avg_hd
                    else:
                        # Too few valid days AND suspiciously low — likely incomplete sheet entry
                        hd = None
                elif eggs > 0 and pop > 0:
                    # Always divide by 7 (full week denominator) so partial-week entries
                    # don't get artificially inflated by a 'days' = 1 denominator
                    hd = (eggs / (pop * 7)) * 100
                    if hd < 15.0:  # < 15% from egg count alone → incomplete data
                        hd = None
                else:
                    hd = None
                
                cum_dep += d['dep_ekor']
                dep_pct_cum = (cum_dep / pop) * 100 if pop > 0 else None
                
                ew_avg = sum(d['ew_vals']) / len(d['ew_vals']) if d['ew_vals'] else None
                em = (ew_avg * eggs / 1000) if (ew_avg and eggs) else None
                
                pakan_g_avg = sum(d['pakan_g_days']) / len(d['pakan_g_days']) if d['pakan_g_days'] else None
                fcr_avg = sum(d['fcr_vals']) / len(d['fcr_vals']) if d['fcr_vals'] else None
                
                results[w] = {
                    'hd': hd,
                    'egg_weight': ew_avg,
                    'egg_mass': em,
                    'pakan_kg': d['pakan_kg'] if d['pakan_kg'] > 0 else None,
                    'pakan_g': pakan_g_avg,
                    'deplesi_ekor': d['dep_ekor'] if d['dep_ekor'] > 0 else None,
                    'deplesi_pct': dep_pct_cum,
                    'fcr': fcr_avg,
                }
            return results
        except Exception as e:
            print(f"Harian Fallback Error: {e}")
            import traceback; traceback.print_exc()
            return None

    def _extract_daily_production_api(self, file_id, populasi):
        print(f"      [DEBUG] Starting daily extraction for file_id: {file_id}")
        try:
            HARIAN_NAMES = ['Data Harian', 'data harian', 'DATA HARIAN']
            rows = None
            for sname in HARIAN_NAMES:
                try:
                    result = self.sheets_service.spreadsheets().values().get(
                        spreadsheetId=file_id,
                        range=f"'{sname}'!A1:Z1500",
                        valueRenderOption='UNFORMATTED_VALUE',
                        dateTimeRenderOption='FORMATTED_STRING'
                    ).execute()
                    candidate = result.get('values', [])
                    if candidate:
                        rows = candidate
                        print(f"      [DEBUG] Found {len(rows)} rows in sheet '{sname}'")
                        break
                except Exception as e:
                    print(f"      [DEBUG] Failed to read '{sname}': {e}")
                    continue

            if not rows:
                print(f"      [DEBUG] No rows found in any Harian sheet.")
                return []

            max_len = max(len(r) for r in rows)
            padded = [r + [None] * (max_len - len(r)) for r in rows]
            df = pd.DataFrame(padded)

            idx = self._find_harian_columns(df)
            if 'date' not in idx or 'week' not in idx or 'btr' not in idx:
                print(f"      [DEBUG] Missing columns in daily extraction. Found: {idx.keys()}")
                return []

            daily_list = []
            pop = populasi or 2000
            current_w = 0
            for i in range(len(df)):
                row = df.iloc[i]
                d_val = row[idx['date']]
                w_val = row[idx['week']]
                
                try:
                    if w_val and str(w_val).strip():
                        current_w = int(float(str(w_val)))
                except: pass

                if current_w <= 0 or current_w > 95: continue
                w = current_w

                date_str = None
                if isinstance(d_val, str) and len(d_val) >= 8:
                    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%y", "%d-%b-%Y"]:
                        try:
                            date_str = datetime.strptime(d_val.split()[0], fmt).strftime("%Y-%m-%d")
                            break
                        except: continue
                if not date_str: continue

                hd_actual = None
                try:
                    if 'hd' in idx:
                        hd_val = self._safe_float(row.iloc[idx['hd']] if hasattr(row, 'iloc') else row[idx['hd']])
                        if hd_val is not None:
                            # if somehow it's less than 1, might be fractional
                            hd_actual = hd_val * 100 if 0 < hd_val <= 1 else hd_val
                    elif 'btr' in idx:
                        btr_raw = row.iloc[idx['btr']] if hasattr(row, 'iloc') else row[idx['btr']]
                        btr = self._safe_float(btr_raw)
                        if btr and btr > 0 and pop > 0:
                            hd_actual = self._clamp((btr / pop) * 100, 5, 2)
                except: pass

                pakan_kg = self._safe_float(row[idx['pakan_kg']]) if 'pakan_kg' in idx else None
                pakan_g = self._safe_float(row[idx['pakan_g']]) if 'pakan_g' in idx else None
                
                deplesi_ekor = None
                if 'dep_ekor' in idx:
                    try: deplesi_ekor = int(float(str(row[idx['dep_ekor']] or 0)))
                    except: pass
                
                fcr = self._safe_float(row[idx['fcr']]) if 'fcr' in idx else None

                if hd_actual is None and deplesi_ekor is None: continue

                daily_list.append({
                    'tanggal': date_str,
                    'usia_minggu': w,
                    'hd_actual': hd_actual,
                    'pakan_kg_hr': self._clamp(pakan_kg, 8, 2),
                    'pakan_gr_ekor': self._clamp(pakan_g, 6, 2),
                    'deplesi_ekor': deplesi_ekor,
                    'fcr_actual': self._clamp(fcr, 6, 3)
                })

            print(f"      [DEBUG] Successfully extracted {len(daily_list)} daily records.")
            return daily_list
        except Exception as e:
            print(f"      [DEBUG] Error extracting daily production api: {e}")
            import traceback
            traceback.print_exc()
            return []

    def get_computed_population(self, file_id, file_name):
        # Strain/farm metadata lookup — used for kandang registration only
        ANCHORS = {
            "KD 4": ("JTP", "PL244T"), "KD 5": ("JTP", "PL244P"),
            "KD 7": ("JTP", "PL244P"), "Jantan": ("JTP", "PL244P"),
            "1A": ("BBK", "AL101"), "KD 2": ("BBK", "AL101"),
            "3A+3B": ("BBK", "AL1001"), "6A": ("BBK", "AL101"),
            "6B": ("BBK", "AL101"), "7A": ("BBK", "AL101"),
            "7B": ("BBK", "AL101"), "9A": ("BBK", "AL101"),
            "9B": ("BBK", "AL101"), "11": ("BBK", "AL101"),
            "12": ("BBK", "AL101"), "14": ("BBK", "AL101"),
            "15": ("BBK", "AL101"), "16": ("BBK", "AL122"),
            "17": ("BBK", "AL122")
        }
        f_norm = file_name.upper().replace(".XLSX", "")
        anchor_key = None
        for k in sorted(ANCHORS.keys(), key=len, reverse=True):
            k_upper = k.upper()
            search_term = re.escape(k_upper.replace("KD ", "").strip())
            if re.search(rf'(\bKD\s*|(?<!\w)){search_term}(\b|(?!\w))', f_norm):
                anchor_key = k; break
        if not anchor_key: return 0, None, None

        farm_type, strain = ANCHORS[anchor_key]
        ceiling_date = datetime.now().date() - timedelta(days=1)  # n-1

        try:
            rows = None
            sheet_name_used = None

            # ── Strictly use Sheets API (Live Data) ──
            HARIAN_NAMES = ['Data Harian', 'data harian', 'DATA HARIAN']
            for sname in HARIAN_NAMES:
                try:
                    result = self.sheets_service.spreadsheets().values().get(
                        spreadsheetId=file_id,
                        range=f"'{sname}'!A1:Z1500"
                    ).execute()
                    candidate = result.get('values', [])
                    if candidate:
                        rows = candidate
                        sheet_name_used = sname
                        print(f"   [POP] Live Sync via Sheets API: '{sname}'")
                        break
                except Exception:
                    continue

            if not rows:
                print(f"   [POP] ERROR: Live Sync failed for {file_name}. Ensure it is a Google Sheet and shared correctly.")
                return 0, None, strain

            # ── Step 1: Find TANGGAL and HIDUP column indices in header rows ──
            col_tgl = None
            col_hidup = None
            header_row_idx = 9

            for r_idx, row in enumerate(rows[:15]):
                upper = [str(v or '').upper().strip() for v in row]
                # Look for Date column keywords
                if any(x in upper for x in ('TANGGAL', 'TGL', 'TGL.', 'DATE')):
                    header_row_idx = r_idx
                    for c_idx, v in enumerate(upper):
                        if v in ('TANGGAL', 'TGL', 'TGL.', 'DATE') and col_tgl is None:
                            col_tgl = c_idx
                    
                    # Search ±2 rows for Population keywords
                    pop_keywords = ('HIDUP', 'POPULASI', 'POP.', 'STOK', 'SISA')
                    for sr in range(max(0, r_idx - 2), min(len(rows), r_idx + 3)):
                        row_content = [str(x or '').upper().strip() for x in rows[sr]]
                        for c_idx, v in enumerate(row_content):
                            if any(k in v for k in pop_keywords) and col_hidup is None:
                                col_hidup = c_idx
                    break

            if col_tgl is None or col_hidup is None:
                print(f"   [POP] Could not find TANGGAL/HIDUP columns in {file_name}")
                return 0, None, strain

            # ── Step 2: Find the latest Hidup value on or before n-1 date ──
            best_pop = None
            best_date = None

            for row in rows[header_row_idx + 1:]:
                if len(row) <= max(col_tgl, col_hidup):
                    continue
                d_val = row[col_tgl]
                row_dt = None

                if isinstance(d_val, datetime):
                    row_dt = d_val.date()
                elif d_val:
                    d_str = str(d_val).strip()
                    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d %H:%M:%S"]:
                        try: 
                            row_dt = datetime.strptime(d_str.split()[0], fmt).date()
                            break
                        except: 
                            continue

                if row_dt and row_dt <= ceiling_date:
                    h_val = row[col_hidup]
                    if h_val is not None and str(h_val).strip():
                        try:
                            pop_int = int(float(str(h_val).replace(",", "").strip()))
                            if pop_int > 0:
                                best_pop = pop_int
                                best_date = row_dt
                        except: continue

            if best_pop is not None:
                # Staleness Check
                days_diff = (datetime.now().date() - best_date).days
                if days_diff > 3:
                    print(f"   [POP] WARNING: Data is STALE ({days_diff} days old). Hidup on {best_date}: {best_pop}")
                else:
                    print(f"   [POP] Hidup on {best_date}: {best_pop} (Freshness: {days_diff} days)")
                
                return best_pop, best_date.strftime("%Y-%m-%d"), strain
            else:
                print(f"   [POP] No Hidup data found up to {ceiling_date}")
                return 0, None, strain

        except Exception as e:
            print(f"   [POP] Error reading {file_name}: {e}")
            return 0, None, strain

    def run_sync(self, root_folder_id, filter_name=None):
        farms = self.get_farm_folders(root_folder_id)
        all_data = []
        for farm in farms:
            farm_name = self._resolve_farm_name(farm['name'])
            files = self.list_xlsx_files(farm['id'])
            for file in files:
                if filter_name and filter_name.upper() not in file['name'].upper():
                    continue
                print(f"Syncing {file['name']} ({farm_name})...")
                pop, date_str, strain = self.get_computed_population(file['id'], file['name'])
                content = self.download_file(file['id'])
                extracted = self.extract_data_from_excel(io.BytesIO(content.getvalue()), farm_name, file['name'], file_id=file['id'], populasi=pop)
                if date_str: extracted['populasi'] = pop; extracted['last_recorded_date'] = date_str
                if strain: extracted['strain'] = strain
                all_data.append(extracted)
        return all_data

    FARM_NAME_MAP = {
        'RECORDING JTP': 'Kandang JTP', 'RECORDING BBK': 'Kandang BBK',
        'REC KANDANG JTP': 'Kandang JTP', 'REC KANDANG BBK': 'Kandang BBK',
        'REC JTP': 'Kandang JTP', 'REC BBK': 'Kandang BBK',
        'JTP': 'Kandang JTP', 'BBK': 'Kandang BBK',
    }

    def _resolve_farm_name(self, raw_name: str) -> str:
        """Map any raw Google Drive folder name to the canonical farm name."""
        upper = raw_name.upper().strip()
        # Exact map lookup first
        if upper in self.FARM_NAME_MAP:
            return self.FARM_NAME_MAP[upper]
        # Substring fallback: if name contains BBK or JTP
        if 'BBK' in upper:
            return 'Kandang BBK'
        if 'JTP' in upper:
            return 'Kandang JTP'
        return raw_name

    def run_sync_multi(self):
        bbk_ids = os.getenv("GOOGLE_DRIVE_BBK_IDS", "").split(",")
        jtp_ids = os.getenv("GOOGLE_DRIVE_JTP_IDS", "").split(",")
        
        all_data = []
        # Process BBK
        for fid in [f for f in bbk_ids if f]:
            for file in self.list_xlsx_files(fid):
                print(f"Syncing {file['name']} (BBK)...")
                pop, date_str, strain = self.get_computed_population(file['id'], file['name'])
                content = self.download_file(file['id'])
                extracted = self.extract_data_from_excel(io.BytesIO(content.getvalue()), "Kandang BBK", file['name'], file_id=file['id'], populasi=pop)
                if date_str: extracted['populasi'] = pop; extracted['last_recorded_date'] = date_str
                if strain: extracted['strain'] = strain
                all_data.append(extracted)
                
        # Process JTP
        for fid in [f for f in jtp_ids if f]:
            for file in self.list_xlsx_files(fid):
                print(f"Syncing {file['name']} (JTP)...")
                pop, date_str, strain = self.get_computed_population(file['id'], file['name'])
                content = self.download_file(file['id'])
                extracted = self.extract_data_from_excel(io.BytesIO(content.getvalue()), "Kandang JTP", file['name'], file_id=file['id'], populasi=pop)
                if date_str: extracted['populasi'] = pop; extracted['last_recorded_date'] = date_str
                if strain: extracted['strain'] = strain
                all_data.append(extracted)
        return all_data

if __name__ == "__main__":
    tool = GoogleDriveTool()
    root = os.getenv("GOOGLE_DRIVE_ROOT_ID")
    if root:
        results = tool.run_sync(root)
        with open('.tmp/ingested_data.json', 'w') as f: json.dump(results, f, indent=2)
