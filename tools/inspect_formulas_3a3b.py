import io
import os
import sys
import openpyxl
from dotenv import load_dotenv

# Add the tools directory to sys.path
sys.path.append('c:\\Users\\ASUS\\OneDrive\\Documents\\Adcanded Poultry Finance Control\\tools')
sys.path.append('c:\\Users\\ASUS\\OneDrive\\Documents\\Adcanded Poultry Finance Control')

from google_drive_tool import GoogleDriveTool

def inspect_formulas_3a3b():
    load_dotenv('c:\\Users\\ASUS\\OneDrive\\Documents\\Adcanded Poultry Finance Control\\.env')
    tool = GoogleDriveTool()
    file_id = '1St1UKPst2EZMpEJK8bT5-ls3bXwus_4n'
    
    content = tool.download_file(file_id)
    # Open with data_only=False to see formulas
    wb = openpyxl.load_workbook(io.BytesIO(content.getvalue()), data_only=False)
    
    if 'Data_Out' in wb.sheetnames:
        ws = wb['Data_Out']
        print("Inspecting Data_Out Row 25 (Week 31) Formulas:")
        # Row 25 is index 25 in 1-indexed Excel? Or 24?
        # My previous inspect said Row 25: Week=31
        row = 25
        for col in range(1, 15):
            cell = ws.cell(row=row, column=col)
            print(f"Col {col}: Value={cell.value}")
    else:
        print("Data_Out sheet not found.")

if __name__ == "__main__":
    inspect_formulas_3a3b()
