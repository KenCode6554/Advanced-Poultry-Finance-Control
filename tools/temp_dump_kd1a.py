import os
import openpyxl
from google_drive_tool import GoogleDriveTool
import dotenv

dotenv.load_dotenv()

def debug():
    tool = GoogleDriveTool()
    root_folder = os.getenv("GOOGLE_DRIVE_ROOT_ID")
        
    farms = tool.get_farm_folders(root_folder)
    
    for farm in farms:
        files = tool.list_xlsx_files(farm['id'])
        for file in files:
            name_upper = file['name'].upper()
            if "1A" in name_upper and "BBK" in name_upper:
                print(f"Found {file['name']}, downloading...")
                content = tool.download_file(file['id'])
                wb = openpyxl.load_workbook(content, data_only=True, read_only=True)
                
                for sheet_name in wb.sheetnames:
                    if "Harian" in sheet_name:
                        ws = wb[sheet_name]
                        print(f"Searching in '{sheet_name}'...")
                        
                        all_rows = list(ws.iter_rows(values_only=True))
                        
                        HIDUP_COL = 4
                        for r_idx in range(6, min(15, len(all_rows))):
                            row = all_rows[r_idx]
                            for c_idx, val in enumerate(row[:15]): 
                                cell_text = str(val or '').strip().lower()
                                if 'hidup' in cell_text and 'awal' not in cell_text and 'total' not in cell_text:
                                    HIDUP_COL = c_idx
                        print(f"Dynamic HIDUP_COL: {HIDUP_COL}")

                        print("\n--- Rows from Row 400 onwards (tail) ---")
                        for r_idx in range(len(all_rows)-30, len(all_rows)):
                            if r_idx >= 0:
                                print(f"Row {r_idx+1}: {all_rows[r_idx][:10]}")

                wb.close()
                return

if __name__ == "__main__":
    debug()
