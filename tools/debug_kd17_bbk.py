"""
debug_kd17_bbk.py
Quick diagnostic to find the 'Hidup' column in Rec P. fajar kd 17 BBK
"""
import os, sys
from dotenv import load_dotenv
from google_drive_tool import GoogleDriveTool

load_dotenv()
tool    = GoogleDriveTool()
root_id = os.getenv("GOOGLE_DRIVE_ROOT_ID")

farms = tool.get_farm_folders(root_id)
target_file = None

for farm in farms:
    files = tool.list_xlsx_files(farm["id"])
    for f in files:
        if "kd 17" in f["name"].lower():
            target_file = f
            break
    if target_file:
        break

if not target_file:
    print("File not found.")
    sys.exit(1)

print(f"File: {target_file['name']}")

import openpyxl
content = tool.download_file(target_file["id"])
wb = openpyxl.load_workbook(content, data_only=True, read_only=True)
print(f"Sheets: {wb.sheetnames}")

for sname in wb.sheetnames:
    if "harian" in sname.lower():
        ws = wb[sname]
        print(f"\nSheet: {sname}")
        print("Rows 1-15:")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
            print(f"  Row {i:2d}: {row[:10]}")
        break

wb.close()
