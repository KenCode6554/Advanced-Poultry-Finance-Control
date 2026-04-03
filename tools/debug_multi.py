"""
debug_multi.py
Diagnose multiple files to see why headers moved or values are missing.
"""
import os, sys
import openpyxl
from dotenv import load_dotenv
from google_drive_tool import GoogleDriveTool

load_dotenv()
tool = GoogleDriveTool()

files_to_check = [
    "Recording BBK Kd 9a AL1001.xlsx",
    "Rec P. fajar kd 17 BBK (AL 1001).xlsx"
]

root_id = os.getenv("GOOGLE_DRIVE_ROOT_ID")
farms = tool.get_farm_folders(root_id)

for fname in files_to_check:
    print(f"\n--- Checking: {fname} ---")
    target_f = None
    for farm in farms:
        fs = tool.list_xlsx_files(farm["id"])
        for f in fs:
            if f["name"] == fname:
                target_f = f
                break
        if target_f: break
    
    if not target_f:
        print("Not found in Drive.")
        continue

    content = tool.download_file(target_f["id"])
    wb = openpyxl.load_workbook(content, data_only=True, read_only=True)
    
    # Try to find 'Data Harian'
    sname = None
    for n in wb.sheetnames:
        if "Data Harian" in n:
            sname = n
            break
    
    if not sname:
        print(f"No Data Harian sheet. Actual: {wb.sheetnames}")
        wb.close()
        continue
        
    ws = wb[sname]
    print(f"Sheet: {sname}")
    
    # Print rows 8-12 to see headers
    print("Header rows (8-12):")
    rows = list(ws.iter_rows(min_row=8, max_row=12, values_only=True))
    for i, r in enumerate(rows, start=8):
        # Print column indices to be clear
        # we want to see where "Hidup" or "Populasi" is
        print(f"  Row {i:2d}: ", end="")
        for c_idx, val in enumerate(r):
            if val is not None:
                print(f"[{c_idx}:{val}] ", end="")
        print()

    # If it's Kd 17, also check the tail
    if "kd 17" in fname.lower():
        print("\nTail of data (check for pre-filled formulas):")
        # Find last row with any value in Col A (Date)
        # iter_rows doesn't give last row easily without reading all
        all_data = list(ws.iter_rows(values_only=True))
        last_idx = len(all_data) - 1
        # Show last 10 rows
        for i in range(max(0, last_idx - 10), last_idx + 1):
            r = all_data[i]
            print(f"  Row {i+1:3d}: {r[:6]}") # Date, Usia, Deplesi, Pop, Afkir, Prod

    wb.close()
