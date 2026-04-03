
import openpyxl
import os
from dotenv import load_dotenv
from google_drive_tool import GoogleDriveTool

load_dotenv()
tool = GoogleDriveTool()

files_to_check = [
    {"name": "Rec P. fajar kd 17 BBK (AL 1001).xlsx", "id": None},
    {"name": "Rec P. fajar kd 3A+3B (AL 1001).xlsx", "id": None}
]

# Find IDs
root_id = os.getenv("GOOGLE_DRIVE_ROOT_ID")
farms = tool.get_farm_folders(root_id)
for farm in farms:
    objs = tool.list_xlsx_files(farm['id'])
    for obj in objs:
        for f in files_to_check:
            if f['name'] == obj['name']:
                f['id'] = obj['id']

for f_info in files_to_check:
    if not f_info['id']:
        print(f"Could not find ID for {f_info['name']}")
        continue
    
    print(f"\n--- Checking {f_info['name']} ---")
    content = tool.download_file(f_info['id'])
    wb = openpyxl.load_workbook(content, data_only=False)
    sheet = None
    for name in wb.sheetnames:
        if 'Data Harian' in name:
            sheet = wb[name]
            break
    
    if not sheet:
        print("Data Harian sheet not found")
        continue
    
    # Look at first 30 rows
    rows = list(sheet.iter_rows(min_row=1, max_row=100, min_col=1, max_col=10, values_only=False))
    for r_idx, row in enumerate(rows[:30]):
        vals = [str(c.value) if c.value is not None else "" for c in row]
        print(f"Row {r_idx+1} | " + " | ".join(vals[:10]))
        
        # Check for formulas in these rows too
        for c_idx, c in enumerate(row):
            f_val = c.value
            if f_val and isinstance(f_val, str) and 'IMPORTRANGE' in f_val:
                print(f"  !!! FOUND FORMULA at Row {r_idx+1} Col {c_idx+1}: {f_val}")
        
    print(f"Finished checking {f_info['name']}")
