import os
import re
import datetime
from google_drive_tool import GoogleDriveTool
from googleapiclient.discovery import build
import dotenv
import openpyxl

dotenv.load_dotenv()

def parse_date(date_val):
    if isinstance(date_val, datetime.datetime):
        return date_val.strftime('%Y-%m-%d')
    if isinstance(date_val, str):
        # basic parsing
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d-%m-%Y', '%Y-%m-%d %H:%M:%S', '%m/%d/%y'):
            try:
                return datetime.datetime.strptime(date_val, fmt).strftime('%Y-%m-%d')
            except ValueError:
                pass
    return None

def debug():
    tool = GoogleDriveTool()
    sheets_service = build('sheets', 'v4', credentials=tool.creds)
    root_folder = os.getenv("GOOGLE_DRIVE_ROOT_ID")
    farms = tool.get_farm_folders(root_folder)
    
    for farm in farms:
        files = tool.list_xlsx_files(farm['id'])
        for file in files:
            name_upper = file['name'].upper()
            if not ("KD" in name_upper and ("BBK" in name_upper or "JTP" in name_upper)):
                continue
                
            print(f"\n--- Processing {file['name']} ---")
            
            content = tool.download_file(file['id'])
            # We need formula strings to parse IMPORTRANGE, so data_only=False
            wb = openpyxl.load_workbook(content, data_only=False, read_only=True)
            
            # Also load data_only=True to get the last valid numbers
            content.seek(0)
            wb_data = openpyxl.load_workbook(content, data_only=True, read_only=True)
            
            target_sheet = None
            for name in wb.sheetnames:
                if 'Harian' in name:
                    target_sheet = name
                    break
                    
            if not target_sheet:
                continue
                
            ws_form = list(wb[target_sheet].iter_rows(values_only=True))
            ws_data = list(wb_data[target_sheet].iter_rows(values_only=True))
            
            HIDUP_COL = 4  # Default
            for r_idx in range(6, min(15, len(ws_data))):
                row = ws_data[r_idx]
                for c_idx, val in enumerate(row[:15]): 
                    cell_text = str(val or '').strip().lower()
                    if 'hidup' in cell_text and 'awal' not in cell_text and 'total' not in cell_text:
                        HIDUP_COL = c_idx

            latest_pop = None
            latest_date_obj = None
            latest_usia = None
            latest_source_row = 0
            
            # 1. Find the last VALID calculated population
            for r_idx in range(8, len(ws_data)):
                row_data = ws_data[r_idx]
                if not row_data or len(row_data) <= HIDUP_COL: continue
                
                p_val = row_data[HIDUP_COL]
                if p_val is None: continue
                
                p_str = str(p_val).strip()
                if not p_str or p_str.startswith('#'): continue
                
                try:
                    pop_val = int(float(p_str.replace(',', '')))
                    if pop_val > 0:
                        latest_pop = pop_val
                        latest_source_row = r_idx
                        if len(row_data) > 0: latest_date_obj = row_data[0]
                        if len(row_data) > 1: latest_usia = row_data[1]
                except (ValueError, TypeError):
                    pass
            
            print(f"Base Valid Pop: {latest_pop} at {parse_date(latest_date_obj)} (Row {latest_source_row+1})")
            
            if not latest_pop:
                print("No base population found.")
                continue
                
            # 2. Check rows after the valid one for IMPORTRANGE
            target_spreadsheet_id = None
            target_sheet_name = None
            
            for r_idx in range(latest_source_row + 1, min(latest_source_row + 15, len(ws_form))):
                row_form = ws_form[r_idx]
                if len(row_form) > 2 and isinstance(row_form[2], str) and 'IMPORTRANGE' in row_form[2]:
                    # Extract URL and sheet
                    # e.g. IMPORTRANGE(""https://docs.google.com/spreadsheets/d/1lg9OARy-pvedq8GLA3D9MVwwleVmQKGQ8geeUxWVxi4/edit?usp=sharing"",""Kandang 1A!C:C"")
                    match = re.search(r'spreadsheets/d/([a-zA-Z0-9_\-]+).*?""([^\",!]+)!', row_form[2])
                    if match:
                        target_spreadsheet_id = match.group(1)
                        target_sheet_name = match.group(2).strip()
                        print(f"Detected IMPORTRANGE -> ID: {target_spreadsheet_id}, Sheet: {target_sheet_name}")
                        break
            
            if target_spreadsheet_id and target_sheet_name:
                # 3. Fetch from source
                try:
                    # Let's get cols A to D
                    res = sheets_service.spreadsheets().values().get(
                        spreadsheetId=target_spreadsheet_id, 
                        range=f"'{target_sheet_name}'!A:D"
                    ).execute()
                    source_rows = res.get('values', [])
                    
                    # Map source dates to deplesi and cull
                    # Row structure might varying, assume Col A is Date, Col C is Deplesi
                    source_deplesi_map = {}
                    for row in source_rows:
                        if not row or len(row) < 3: continue
                        d_str = row[0]
                        dep_val = row[2]
                        if not d_str or not dep_val: continue
                        try:
                            # Basic date parse
                            for fmt in ('%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d'):
                                try:
                                    dt = datetime.datetime.strptime(d_str, fmt).strftime('%Y-%m-%d')
                                    source_deplesi_map[dt] = int(str(dep_val).replace(',', ''))
                                    break
                                except ValueError:
                                    pass
                        except Exception:
                            pass
                    
                    # 4. Patch the gaps!
                    current_pop = latest_pop
                    final_date = parse_date(latest_date_obj)
                    
                    for r_idx in range(latest_source_row + 1, len(ws_data)):
                        row_data = ws_data[r_idx]
                        if not row_data or len(row_data) == 0: continue
                        
                        target_date_str = parse_date(row_data[0])
                        if not target_date_str: break
                        
                        inc_deplesi = source_deplesi_map.get(target_date_str, None)
                        if inc_deplesi is not None:
                            current_pop -= inc_deplesi
                            final_date = target_date_str
                            print(f"Patched {target_date_str}: Deplesi {inc_deplesi} -> New Pop: {current_pop}")
                        else:
                            # No more data in source sheet
                            break
                            
                    print(f"*** FINAL PATCHED RESULT: {current_pop} on {final_date} ***")
                    
                except Exception as e:
                    print(f"Failed to fetch source API: {e}")
            else:
                print("No IMPORTRANGE formula found, using base pop.")
                
            wb.close()
            wb_data.close()

if __name__ == "__main__":
    debug()
