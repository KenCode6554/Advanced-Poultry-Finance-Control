
import os
import openpyxl
from google_drive_tool import GoogleDriveTool
from dotenv import load_dotenv

load_dotenv()
tool = GoogleDriveTool()

target_file = "Rec P. fajar kd 3A+3B (AL 1001).xlsx"
root_id = os.getenv("GOOGLE_DRIVE_ROOT_ID")
folders = tool.get_farm_folders(root_id)

found_file = False
for folder in folders:
    files = tool.list_xlsx_files(folder['id'])
    for f in files:
        if f['name'] == target_file:
            print(f"Found {target_file}. Downloading...")
            content = tool.download_file(f['id'])
            wb = openpyxl.load_workbook(content, data_only=False)
            found_formula = False
            for sname in wb.sheetnames:
                sheet = wb[sname]
                print(f"Checking sheet: {sname}")
                # Check first 50x50 cells for formula
                for r_idx, row in enumerate(sheet.iter_rows(max_row=50, max_col=20)):
                    for c_idx, cell in enumerate(row):
                        val = str(cell.value) if cell.value else ""
                        if 'IMPORTRANGE' in val:
                            print(f"  FOUND FORMULA at {sname}!{cell.coordinate}: {val}")
                            found_formula = True
            if not found_formula:
                print("  NO IMPORTRANGE FOUND in first 50x20 of any sheet.")
            found_file = True
            break
    if found_file: break
