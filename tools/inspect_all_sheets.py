"""
inspect_all_sheets.py
Print first 10 rows of all sheets in KD 5 JTP.
"""
import os
import openpyxl
from google_drive_tool import GoogleDriveTool

tool = GoogleDriveTool()
root_id = os.getenv("GOOGLE_DRIVE_ROOT_ID")
folders = tool.get_farm_folders(root_id)
jtp_folder = next(f for f in folders if 'JTP' in f['name'].upper())
files = tool.list_xlsx_files(jtp_folder['id'])

target = 'REC KD 5 PL241P JTP Mojogedang .xlsx'
file_id = next(f['id'] for f in files if f['name'] == target)

content_io = tool.download_file(file_id)
wb = openpyxl.load_workbook(content_io, data_only=True, read_only=True)

for sheet_name in wb.sheetnames:
    print(f"\n--- Sheet: {sheet_name} ---")
    sheet = wb[sheet_name]
    for r in range(1, 15):
        row = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
        print(f"Row {r:2}: {row}")

    # Also check the end of the sheet
    print("Tail:")
    for r in range(390, 420):
        row = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
        if any(v is not None for v in row):
            print(f"Row {r:3}: {row}")
