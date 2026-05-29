import sys, os, io, datetime
sys.path.append(os.getcwd())
from tools.google_drive_tool import GoogleDriveTool
import openpyxl

tool = GoogleDriveTool()
fid = '1IaVF9iIXFfREX1V6d6yzPugYhVoSfEwF'
content = tool.download_file(fid)
wb = openpyxl.load_workbook(io.BytesIO(content.getvalue()), data_only=True)
harian = next((wb[n] for n in wb.sheetnames if 'HARIAN' in n.upper()), None)
sheet = harian or wb.worksheets[0]

target = datetime.date(2026, 4, 11)

# Find April 11 row
for r in range(10, sheet.max_row + 1):
    d = sheet.cell(row=r, column=1).value
    if isinstance(d, datetime.datetime) and d.date() == target:
        print(f'Row {r} = April 11:')
        for c in range(1, 12):
            v = sheet.cell(row=r, column=c).value
            print(f'  Col {chr(64+c)} ({c}): {v}')
        break

# Also print last ~5 rows with dates
print('\nLast 5 date rows:')
last_rows = []
for r in range(10, sheet.max_row + 1):
    d = sheet.cell(row=r, column=1).value
    if isinstance(d, datetime.datetime):
        last_rows.append((r, d.date(), sheet.cell(row=r, column=5).value))
for row in last_rows[-5:]:
    print(f'  Row {row[0]} | {row[1]} | ColE(Hidup)={row[2]}')
