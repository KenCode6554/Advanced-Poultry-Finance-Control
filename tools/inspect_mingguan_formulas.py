import io
import os
import sys
import openpyxl
from dotenv import load_dotenv

# Add the tools directory to sys.path
sys.path.append('c:\\Users\\ASUS\\OneDrive\\Documents\\Adcanded Poultry Finance Control\\tools')
sys.path.append('c:\\Users\\ASUS\\OneDrive\\Documents\\Adcanded Poultry Finance Control')

from google_drive_tool import GoogleDriveTool

def inspect_mingguan_formulas():
    load_dotenv('c:\\Users\\ASUS\\OneDrive\\Documents\\Adcanded Poultry Finance Control\\.env')
    tool = GoogleDriveTool()
    file_id = '1St1UKPst2EZMpEJK8bT5-ls3bXwus_4n'
    
    content = tool.download_file(file_id)
    wb = openpyxl.load_workbook(io.BytesIO(content.getvalue()), data_only=False)
    
    if 'Data Mingguan' in wb.sheetnames:
        ws = wb['Data Mingguan']
        print("Inspecting Data Mingguan Row 24 (Week 31) Formulas:")
        row = 24
        # Check first 20 columns
        for col in range(1, 21):
            cell = ws.cell(row=row, column=col)
            print(f"Col {col}: Value={cell.value}")
    else:
        print("Data Mingguan sheet not found.")

if __name__ == "__main__":
    inspect_mingguan_formulas()
